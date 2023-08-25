import xml.etree.ElementTree as ET
import sqlite3
import os
import shutil
from openpyxl import load_workbook

days = {
    d: i
    for i, d in enumerate(["100000", "010000", "001000", "000100", "000010", "000001"])
}

full_query = """
select subject.name, teacher.name, class.name, `group`.name, classroom.name, card.period, card.days 
from card
join classroom on card.classroomid=classroom.id
join lesson on lesson.id=card.lessonid
join lesson_to_teacher on lesson.id=lesson_to_teacher.lessonid
join teacher on teacher.id=lesson_to_teacher.teacherid
join subject on subject.id=lesson.subjectid
join class on lesson.classid=class.id
join `group` on lesson.groupid=`group`.id
order by teacher.name
;
"""
short = dict(
    [
        ("Алгебра", "Алг"),
        ("Алгебра и начала анализа", "Алг"),
        ("БЕСЕДЫ КЛ РУК", "Беседа"),
        ("Биология", "Биол"),
        ("Введение в физику", "Физика"),
        ("География", "Геогр"),
        ("Геометрия", "Геом"),
        ("ИКТ и робототехника", "Робот."),
        ("Индивидуальный прект", "Проект"),
        ("Иностранный язык", "Ин.яз"),
        ("Информатика и ИКТ", "Инф"),
        ("Искусство", "Иск-во"),
        ("Литература", "Лит-ра"),
        ("История", "История"),
        ("Математика", "Матем"),
        ("ОБЖ", "ОБЖ"),
        ("Обществознание", "Общ"),
        ("ОДНКР Санкт-Петербург – хранитель духовных традиций России", "ОДНКР"),
        ("Русский язык", "Рус.яз"),
        ("Физика", "Физика"),
        ("Физическая культура", "Физ-ра"),
        ("Химия", "Химия"),
        ("", ""),
    ]
)


def make_output(filename="output.xlsx"):
    wb = load_workbook("template.xlsx")
    with sqlite3.connect(".cache.db") as db:
        teachers = wb["Учителя"]
        make_teachers(db, teachers)
        students5 = wb["Ученики5"]
        make_students5(db, students5)
        students = wb["Ученики"]
        make_students(db, students)
    wb.save(filename)


def make_teachers(db, teachers):
    bx, by = 2, 3
    last = None
    for _, name, klass, _, _, time, day in db.execute(full_query):
        if name != last:
            by += 1
            _name = name.split()
            _name = f"{_name[0]} {_name[1][0]}.{_name[2][0]}."
            teachers.cell(by, 1, _name)
        last = name
        time = int(time)
        x = bx + 8 * days[day] + time
        teachers.cell(by, x, klass)


def make_students5(db, students5):
    by = 3
    classes = [
        "5-1",
        "5-2",
        "6-1",
        "6-2",
        "7-1",
        "7-2",
        "8-1",
        "8-2",
        "8-3",
        "8-4",
        "9-1",
        "9-2",
        "9-3",
        "9-4",
        "9-5",
        "9-6",
        "10-1",
        "10-2",
        "10-3",
        "10-4",
        "10-5",
        "10-6",
        "10-7",
        "10-8",
        "11-1",
        "11-3",
        "11-4",
        "11-5",
        "11-6",
        "11-7",
        "11-8",
    ]
    _sheets = []
    for x in [5, 5, 5, 5, 6, 6]:
        _sheets.append(classes[:x])
        classes = classes[x:]
    sheets = {
        klass: x
        for x, classes in zip([5, 29, 53, 77, 101, 129], _sheets)
        for klass in classes
    }
    pos = {klass: i for classes in _sheets for i, klass in enumerate(classes)}
    for subject, _, klass, group, room, time, day in db.execute(full_query):
        subject = short[subject]
        time = int(time)
        x = sheets[klass] + pos[klass] * 4
        y = by + days[day] * 8 + time
        if students5.cell(y, x).value != None:
            x += 2
        students5.cell(y, x, subject)
        students5.cell(y, x + 1, room)


def make_students(db, students):
    bx, by = 4, 3
    classes = [
        "5-1",
        "5-2",
        "6-1",
        "6-2",
        "7-1",
        "7-2",
        "8-1",
        "8-2",
        "8-3",
        "8-4",
        "9-1",
        "9-2",
        "9-3",
        "9-4",
        "9-5",
        "9-6",
        "10-1",
        "10-2",
        "10-3",
        "10-4",
        "10-5",
        "10-6",
        "10-7",
        "10-8",
        "11-1",
        "11-3",
        "11-4",
        "11-5",
        "11-6",
        "11-7",
        "11-8",
    ]
    classes = {k: i for i, k in enumerate(classes)}
    for subject, _, klass, group, room, time, day in db.execute(full_query):
        subject = short[subject]
        time = int(time)
        klass = classes[klass]
        x = bx + klass // 4 * 20 + klass % 4 * 4
        y = by + days[day] * 8 + time
        if students.cell(y, x).value != None:
            x += 2
        students.cell(y, x, subject)
        students.cell(y, x + 1, room)


def create_cache(xml_in, cache_name=".cache.db"):
    if os.path.exists(cache_name):
        os.remove(cache_name)
    tree = ET.parse(xml_in)
    with sqlite3.connect(cache_name) as db:
        for e in tree.getroot():
            if e.tag == "cards":
                db.execute("CREATE TABLE card (lessonid, period, days, classroomid);")
                for t in e:
                    d = t.attrib
                    x = (d["lessonid"], d["period"], d["days"], d["classroomids"])
                    db.execute("INSERT INTO card VALUES (?, ?, ?, ?);", x)
            elif e.tag == "lessons":
                db.execute("CREATE TABLE lesson (id, subjectid, classid, groupid);")
                db.execute("CREATE TABLE lesson_to_teacher (lessonid, teacherid);")
                for t in e:
                    d = t.attrib
                    x = (d["id"], d["subjectid"], d["classids"], d["groupids"])
                    for teacher in d["teacherids"].split(","):
                        db.execute(
                            "INSERT INTO lesson_to_teacher VALUES (?, ?);",
                            [d["id"], teacher],
                        )
                    db.execute("INSERT INTO lesson VALUES (?, ?, ?, ?);", x)
            elif e.tag == "classrooms":
                db.execute("CREATE TABLE classroom (id, name);")
                for t in e:
                    d = t.attrib
                    x = (d["id"], d["short"])
                    db.execute("INSERT INTO classroom VALUES (?, ?);", x)
            elif e.tag == "classes":
                db.execute("CREATE TABLE class (id, name);")
                for t in e:
                    d = t.attrib
                    x = (d["id"], d["short"])
                    db.execute("INSERT INTO class VALUES (?, ?);", x)
            elif e.tag == "subjects":
                db.execute("CREATE TABLE subject (id, name);")
                for t in e:
                    d = t.attrib
                    x = (d["id"], d["name"])
                    db.execute("INSERT INTO subject VALUES (?, ?);", x)
            elif e.tag == "teachers":
                db.execute("CREATE TABLE teacher (id, name);")
                for t in e:
                    d = t.attrib
                    x = (d["id"], d["name"])
                    db.execute("INSERT INTO teacher VALUES (?, ?);", x)
            elif e.tag == "groups":
                db.execute("CREATE TABLE `group` (id, classid, name);")
                for t in e:
                    d = t.attrib
                    x = (d["id"], d["classid"], d["name"])
                    db.execute("INSERT INTO `group` VALUES (?, ?, ?);", x)
