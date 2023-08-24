from openpyxl import load_workbook
from Item import *
from TimeTable import *
import xlsxwriter
import tkinter as tk
from tkinter.filedialog import askdirectory, askopenfilename


def setOutputPath(outpuPath):
    file = askdirectory(".") + "/output.xlsx"
    print(f"{file=}")
    outputPath.set(file)


def getFile(path, outputPath):
    file = askopenfilename(parent=root, title="Choose a xlsx file", initialdir=".")
    path.set(file)
    p = file
    name = list(p.split("/"))[-1][:-4]
    if len(list(p.split("/"))) == 1:
        name = list(p.split("\\"))[-1][:-4]
    childes = list(p.split("/"))[:-1]
    isWind = False
    print(f"{childes=}")
    if len(childes) == 0:
        childes = list(p.split("\\"))[:-1]
        isWind = True
    try:
        if childes[0][-1] == ":":
            isWind = True
    except Exception:
        pass
    if childes[0] == "":
        childes = childes[1:]
    op = ""
    for child in childes:
        op += "/" + child
    if isWind:
        op = op[1:]
    op += "/output.xlsx"
    outputPath.set(op)


def box(work_book, work_sheet, first_row, first_col, rows_count, cols_count):
    # top left corner
    work_sheet.conditional_format(
        first_row,
        first_col,
        first_row,
        first_col,
        {
            "type": "formula",
            "criteria": "True",
            "format": work_book.add_format({"top": 2, "left": 2}),
        },
    )
    # top right corner
    work_sheet.conditional_format(
        first_row,
        first_col + cols_count - 1,
        first_row,
        first_col + cols_count - 1,
        {
            "type": "formula",
            "criteria": "True",
            "format": work_book.add_format({"top": 2, "right": 2}),
        },
    )
    # bottom left corner
    work_sheet.conditional_format(
        first_row + rows_count - 1,
        first_col,
        first_row + rows_count - 1,
        first_col,
        {
            "type": "formula",
            "criteria": "True",
            "format": work_book.add_format({"bottom": 2, "left": 2}),
        },
    )
    # bottom right corner
    work_sheet.conditional_format(
        first_row + rows_count - 1,
        first_col + cols_count - 1,
        first_row + rows_count - 1,
        first_col + cols_count - 1,
        {
            "type": "formula",
            "criteria": "True",
            "format": work_book.add_format({"bottom": 2, "right": 2}),
        },
    )

    # top
    work_sheet.conditional_format(
        first_row,
        first_col + 1,
        first_row,
        first_col + cols_count - 2,
        {
            "type": "formula",
            "criteria": "True",
            "format": work_book.add_format({"top": 2}),
        },
    )
    # left
    work_sheet.conditional_format(
        first_row + 1,
        first_col,
        first_row + rows_count - 2,
        first_col,
        {
            "type": "formula",
            "criteria": "True",
            "format": work_book.add_format({"left": 2}),
        },
    )
    # bottom
    work_sheet.conditional_format(
        first_row + rows_count - 1,
        first_col + 1,
        first_row + rows_count - 1,
        first_col + cols_count - 2,
        {
            "type": "formula",
            "criteria": "True",
            "format": work_book.add_format({"bottom": 2}),
        },
    )
    # right
    work_sheet.conditional_format(
        first_row + 1,
        first_col + cols_count - 1,
        first_row + rows_count - 2,
        first_col + cols_count - 1,
        {
            "type": "formula",
            "criteria": "True",
            "format": work_book.add_format({"right": 2}),
        },
    )


def import_timetable(input):
    #  Import timetable
    workbook = load_workbook(input)
    first_sheet = workbook.sheetnames[0]
    # worksheet = workbook["classrooms (classes)"]
    worksheet = workbook[first_sheet]

    print(f"{worksheet.cell(row=4, column=1).value=}")
    # assert worksheet.cell(row=3, column=2).value=="Понед."

    items = []

    # Parse input file
    for student_class in range(5, 260, 8):
        name = worksheet.cell(row=student_class, column=1).value
        for day in range(0, 6):
            for lesson_number in range(8):
                lesson = []
                for param in range(8):
                    cell = worksheet.cell(
                        row=student_class + param, column=2 + day * 8 + lesson_number
                    ).value
                    if cell is not None and cell != "":
                        lesson.append(cell)
                if len(lesson) == 4:
                    items.append(
                        Item(
                            student_class=name,
                            teacher=lesson[1],
                            subject=lesson[0],
                            group=1,
                            day=day,
                            auditory=lesson[3],
                            lesson_number=lesson_number,
                        )
                    )
                elif len(lesson) == 8:
                    items.append(
                        Item(
                            student_class=name,
                            teacher=lesson[1],
                            subject=lesson[0],
                            group=1,
                            day=day,
                            auditory=lesson[3],
                            lesson_number=lesson_number,
                        )
                    )
                    items.append(
                        Item(
                            student_class=name,
                            teacher=lesson[5],
                            subject=lesson[4],
                            group=2,
                            day=day,
                            auditory=lesson[7],
                            lesson_number=lesson_number,
                        )
                    )

    timetable = TimeTable()

    # Create timetable
    for item in items:
        print(f"{item=}")
        if item.student_class not in timetable.classes:
            timetable.classes.append(item.student_class)
            timetable.classes_timetable[item.student_class] = []
        lesson = Lesson(item.subject, item.lesson_number, item.group - 1, item.auditory)
        if len(timetable.classes_timetable[item.student_class]) <= item.day:
            timetable.classes_timetable[item.student_class].append([lesson])
        else:
            timetable.classes_timetable[item.student_class][item.day].append(lesson)
    return timetable, items


def start(path, outputPath):
    input, output = path.get(), outputPath.get()

    timetable, items = import_timetable(input)

    # Create excel file

    workbook = xlsxwriter.Workbook(output)

    students_worksheet = workbook.add_worksheet("students")
    merge_format = workbook.add_format(
        {"bold": 1, "border": 2, "align": "center", "valign": "vcenter"}
    )

    cell_format = workbook.add_format({"bold": 1, "border": 1})

    start_x = 4
    start_y = 2
    lessons = 8
    width = 4

    students_worksheet.set_column(start_x - 1, start_x - 1, 12)
    students_worksheet.set_column(start_x - 2, start_x - 2, 2)

    ind = 0
    for st_class in timetable.classes:
        for day in range(6):
            box(
                workbook,
                students_worksheet,
                start_y + day * lessons,
                start_x + ind * width,
                lessons,
                width,
            )
            try:
                timetable.classes_timetable[st_class][day]
            except:
                print(f"{timetable.classes_timetable=}")
            for lesson in timetable.classes_timetable[st_class][day]:
                students_worksheet.write(
                    start_y + day * lessons + lesson.number,
                    start_x + ind * width + lesson.group * 2,
                    lesson.name,
                )
                students_worksheet.write(
                    start_y + day * lessons + lesson.number,
                    start_x + ind * width + lesson.group * 2 + 1,
                    lesson.auditory,
                )
            # Add classes names in header
            try:
                students_worksheet.merge_range(
                    start_y - 1,
                    start_x + ind * width,
                    start_y - 1,
                    start_x + ind * width + 3,
                    st_class,
                    merge_format,
                )
            except Exception as e:
                print(e)
        ind += 1

    for day in range(0, 6):
        for lesson in range(lessons):
            box(
                workbook,
                students_worksheet,
                start_y + day * lessons,
                start_x - 2,
                lessons,
                2,
            )
            students_worksheet.write(
                start_y + day * lessons + lesson, start_x - 2, str(lesson)
            )
            students_worksheet.write(
                start_y + day * lessons + lesson, start_x - 1, getLessonsTime(lesson)
            )
        students_worksheet.merge_range(
            start_y + day * lessons,
            start_x - 3,
            start_y + day * lessons + lessons - 1,
            start_x - 3,
            getDay(day),
            merge_format,
        )

    # teachers
    for item in items:
        for teacher in item.teachers:
            if teacher not in timetable.teachers:
                timetable.teachers.append(teacher)
                timetable.teachers_timetable[teacher] = []
            lesson = TeacherLesson(item.student_class, item.lesson_number, item.day)
            lesson.classroom = item.auditory
            timetable.teachers_timetable[teacher].append(lesson)

    teachers_worksheet = workbook.add_worksheet("teachers")
    teachers_worksheet.set_column(0, 0, 18)
    teachers_worksheet.set_column(1, 50, 5)

    start_x = 1
    start_y = 3
    lessons = 8

    ind = 0
    for teacher in sorted(timetable.teachers):
        teachers_worksheet.write(start_y + ind, start_x - 1, teacher)
        for lesson in timetable.teachers_timetable[teacher]:
            class_name = lesson.class_name
            if class_name == "10-8интернат":
                class_name = "10-8"
            teachers_worksheet.write(
                start_y + ind,
                start_x + lesson.day * lessons + lesson.number,
                class_name,
            )
        ind += 1

    for day in range(6):
        box(
            workbook,
            teachers_worksheet,
            start_y,
            start_x + day * lessons,
            len(timetable.teachers),
            lessons,
        )
        box(
            workbook,
            teachers_worksheet,
            start_y - 2,
            start_x + day * lessons,
            2,
            lessons,
        )
        teachers_worksheet.merge_range(
            start_y - 2,
            start_x + day * lessons,
            start_y - 2,
            start_x + day * lessons + lessons - 1,
            getDayFull(day),
            merge_format,
        )
        for lesson in range(lessons):
            teachers_worksheet.write(
                start_y - 1, start_x + day * lessons + lesson, lesson
            )

    # teachers / classrooms
    teachers_classrooms_worksheet = workbook.add_worksheet("teachers (classrooms)")
    teachers_classrooms_worksheet.set_column(0, 0, 18)
    teachers_classrooms_worksheet.set_column(1, 50, 5)

    start_x = 1
    start_y = 3
    lessons = 8

    ind = 0
    for teacher in sorted(timetable.teachers):
        teachers_classrooms_worksheet.write(start_y + ind, start_x - 1, teacher)
        for lesson in timetable.teachers_timetable[teacher]:
            classroom = lesson.classroom
            teachers_classrooms_worksheet.write(
                start_y + ind, start_x + lesson.day * lessons + lesson.number, classroom
            )
        ind += 1

    for day in range(6):
        box(
            workbook,
            teachers_classrooms_worksheet,
            start_y,
            start_x + day * lessons,
            len(timetable.teachers),
            lessons,
        )
        box(
            workbook,
            teachers_classrooms_worksheet,
            start_y - 2,
            start_x + day * lessons,
            2,
            lessons,
        )
        teachers_classrooms_worksheet.merge_range(
            start_y - 2,
            start_x + day * lessons,
            start_y - 2,
            start_x + day * lessons + lessons - 1,
            getDayFull(day),
            merge_format,
        )
        for lesson in range(lessons):
            teachers_classrooms_worksheet.write(
                start_y - 1, start_x + day * lessons + lesson, lesson
            )

    # classrooms / class
    for item in items:
        classroom = item.auditory
        if classroom not in timetable.classrooms:
            timetable.classrooms.append(classroom)
            timetable.classrooms_timetable[classroom] = []
        lesson = TeacherLesson(item.student_class, item.lesson_number, item.day)
        timetable.classrooms_timetable[classroom].append(lesson)

    classrooms_worksheet = workbook.add_worksheet("classrooms (classes)")
    classrooms_worksheet.set_column(0, 0, 8)
    classrooms_worksheet.set_column(1, 50, 5)

    start_x = 1
    start_y = 3
    lessons = 8

    ind = 0
    for classroom in sorted(timetable.classrooms):
        classrooms_worksheet.write(start_y + ind, start_x - 1, classroom)
        for lesson in timetable.classrooms_timetable[classroom]:
            class_name = lesson.class_name
            if class_name == "10-8интернат":
                class_name = "10-8"
            classrooms_worksheet.write(
                start_y + ind,
                start_x + lesson.day * lessons + lesson.number,
                class_name,
            )
        ind += 1

    for day in range(6):
        box(
            workbook,
            classrooms_worksheet,
            start_y,
            start_x + day * lessons,
            len(timetable.classrooms),
            lessons,
        )
        box(
            workbook,
            classrooms_worksheet,
            start_y - 2,
            start_x + day * lessons,
            2,
            lessons,
        )
        classrooms_worksheet.merge_range(
            start_y - 2,
            start_x + day * lessons,
            start_y - 2,
            start_x + day * lessons + lessons - 1,
            getDayFull(day),
            merge_format,
        )
        for lesson in range(lessons):
            classrooms_worksheet.write(
                start_y - 1, start_x + day * lessons + lesson, lesson
            )

    print("done!")
    workbook.close()


if __name__ == "__main__":
    # create window
    root = tk.Tk()
    root.geometry("600x400")
    root.title("Summary Parser")

    # path to file
    path = tk.StringVar()
    path.set("...")

    outputPath = tk.StringVar()
    outputPath.set("...")

    pathLabel = tk.Label(text="input")
    pathLabel.grid(row=0, column=0)

    pathEntry = tk.Entry(width=45, textvariable=path)
    pathEntry.grid(row=0, column=1)

    browseButton = tk.Button(text="Browse", command=lambda: getFile(path, outputPath))
    browseButton.grid(row=0, column=2)

    outputPathLabel = tk.Label(text="output")
    outputPathLabel.grid(row=1, column=0)

    outputPathEntry = tk.Entry(width=45, textvariable=outputPath)
    outputPathEntry.grid(row=1, column=1)

    outputPathBrowseButton = tk.Button(text="Browse", command=lambda: setOutputPath(outputPath))
    outputPathBrowseButton.grid(row=1, column=2)

    startButton = tk.Button(text="Start", command=lambda: start(path, outputPath))
    startButton.grid(row=3, column=2)

    root.mainloop()
